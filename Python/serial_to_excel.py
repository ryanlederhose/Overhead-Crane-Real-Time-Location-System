'''
This python script will read data over the specified serial port
at baudrate and parse the data to store into an excel spreadsheet
    @param -b baudrate
    @param -p com port
    @param -e excel spreadsheet
'''
import serial
from datetime import datetime
import openpyxl
import numpy as np
from sklearn.linear_model import LinearRegression
from sklearn.neighbors import KNeighborsRegressor
import sys
import pandas as pd

MAX_READS = 20

'''
    @brief entry point into program
'''
def main():

    global ser
    global client
    global spreadsheet

    #Get com port and baudrate
    com_port = ""
    baudrate = 0
    for i in range(len(sys.argv)):
        if sys.argv[i] == '-b':
            baudrate = int(sys.argv[i + 1])
        elif sys.argv[i] == '-p':
            com_port = sys.argv[i + 1]
        elif sys.argv[i] == '-e':
            spreadsheet = 'Training Data/'  + sys.argv[i + 1] + '.xlsx'

    #Open serial port
    try:
        ser = serial.Serial(com_port, baudrate = baudrate, timeout = 1)
        print("Serial port has been detected")
    except serial.SerialException:
        print("COULD NOT OPEN SERIAL PORT")
        return -1

    #create excel spreadsheet
    try:
        wb = openpyxl.load_workbook(spreadsheet)
        wb.close()
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        create_sheet('Crane 3', wb)
        wb.save(filename=spreadsheet)
        wb.close()

    read_serial()   #start reading serial

''' 
    @brief read data over the serial port to store in an excel spreadsheet
'''
def read_serial():

    #Global variables
    global ser
    global client
    global spreadsheet

    #Local variables
    readFlag = False
    rxBuffer = ""
    craneID = 0
    waitingFlag = 0
    rawAdc = 0
    weightList = []
    reads = 0

    #Create linear regression model
    training_data = pd.read_csv("Training Data/training_data.csv")  #read data
    X_train = training_data.iloc[:, :-1]
    y_train = training_data.iloc[:, -1]
    model = LinearRegression()
    model.fit(X_train, y_train)

    #Enter loop
    while True:
        
        #Try to read serial port
        try:
            c = ser.read().decode('utf-8')
        except serial.SerialException:
            continue
        except UnicodeDecodeError:
            continue
        
        #check for new line character
        if c == '\n':
            readFlag = True 
        elif c == '\r':
            #ignore, prevent returning early
            continue
        else:
            rxBuffer += c #add received char to buffer
            continue
        
        #Check if message is ready to be parsed
        if readFlag == True:
            readFlag = False
            now = datetime.now()    #get time
            
            dataList = rxBuffer.split(" ")
            if (len(dataList) < 4):
                for i in range(len(dataList)):
                    if (dataList[i])[0] == 'i':
                        craneID = (dataList[i])[1::]
                    elif (dataList[i])[0] == 'k':
                        posX = -1
                        posY = -1
                        waitingFlag = 1
                        continue
                if (waitingFlag == 0):                        
                    continue    #incomplete packet, ignore

            #extract variables from message
            for i in range(len(dataList)):
                if (dataList[i])[0] == '\x16':
                    craneID = (dataList[i])[2:3:]
                elif (dataList[i])[0] == 'm':
                    rawAdc = (dataList[i])[1::]
                elif (dataList[i])[0] == 'x':
                    posX = (dataList[i])[1::]
                elif (dataList[i])[0] == 'y':
                    posY = (dataList[i])[1::]
                else:
                    continue

            #open excel sheet    
            wb = openpyxl.load_workbook(spreadsheet)
            crane3Worksheet = wb.__getitem__("Crane 3")

            #predict the mass of the coil based on the adc reading
            trueMass = model.predict(pd.DataFrame({'adc': [int(rawAdc)]}))

            #put the calculated mass into a buffer of latest fifteen values
            weightList.insert(0, float(trueMass))
            if (len(weightList) > 15):
                weightList.pop(len(weightList) - 1)

            #get the average of the mass
            averageTrueMass = 0
            for i in range(len(weightList)):
                averageTrueMass += weightList[i]
            averageTrueMass = float(averageTrueMass / len(weightList))

            #append data
            excelList = [str(now), int(rawAdc), float(trueMass), int(posX), int(posY)]
            crane3Worksheet.append(excelList)
            reads = reads + 1
            
            #save and close file
            wb.save(filename=spreadsheet)      
            wb.close()

            rxBuffer = ""   #reset buffer

            print(rawAdc, trueMass)

            if reads == MAX_READS:
                return
            
'''
    @brief create the excel spreadsheet in which to store data into
    @param titlename title of spreadsheet
    @param wb workbook object of excel doc
'''
def create_sheet(titlename, wb):
    wb.create_sheet(title=titlename)
    sheet = wb.__getitem__(titlename)
    sheet.append(['Time', 'Raw ADC', 'Mass', 'X Position', 'Y Position'])

# run application
if __name__ == "__main__":
    main()